import streamlit as st
import pandas as pd
import plotly.express as px
from PIL import Image
import xlwt
from openpyxl import load_workbook, workbook
from openpyxl.utils import get_column_letter
# 网页名称、标题、子标题
def head_set():
    st.set_page_config(page_title='质差用户数据分析系统')
    st.header('质差用户数据分析可视化展示')

# 上传分析文件
def excel_upload():
    global uploaded_file  # 全局变量
    uploaded_file = st.file_uploader("上传分析文件",
                                     type="xlsx")

def shuju():
    df = pd.read_excel(uploaded_file, header=0, usecols=[1, 3, 4,14,38, 41, 46, 61])# 读取上传文件
    def fun(df):
        ci = {}
        for i in df.columns:
            ci[i] = df[i].value_counts()
        return ci
    dd = fun(df)
    f = xlwt.Workbook()  # 创建工作薄
    she1 = f.add_sheet(u'sheet1', cell_overwrite_ok=True)  # 创建sheet
    pa = xlwt.Pattern()
    pa.pattern = xlwt.Pattern.SOLID_PATTERN
    pa.pattern_fore_colour = 5
    style = xlwt.XFStyle()
    style.pattern = pa
    ai = xlwt.Alignment()
    ai.horz = 0x02  # 设置水平居中
    ai.vert = 0x01  # 设置垂直居中
    style.alignment = ai
    # 获取字典的键
    list1 = [k for k in dd]
    a = 0
    b = 0
    for s in range(len(dd)):
        b = a + 1
        # 写入第一行
        # sheet1.write_merge(0, 1, k, l, list_[s], style)
        # sheet1.write_merge(0, 1, k, l, list_[s], style)
        she1.write(0, a, list1[s])
        she1.write(0, b, list1[s] + "重复次数")
        # 写入内容
        j = 1
        for c, d in zip(dd[list1[s]], dd[list1[s]].index):
            she1.write(j, a, d)  # 循环写入 竖着写
            she1.write(j, b, c)  # 循环写入 竖着写
            j = j + 1
        a = a + 2
    f.save('统计1.xlsx')
    excel_da = '统计1.xlsx'
    da = pd.read_excel("统计1.xlsx")
    sheet_name1 = 'sheet1'
    # 筛选满足要求的数据
    exectop5 = da[(da['机顶盒业务账号重复次数'] >= 3) & (da['olt设备ip重复次数'] >= 20)]
    #cess = da[da['olt设备ip重复次数'] >= 20]
    ba = exectop5.to_excel('筛选数据后的文件2.xlsx', sheet_name='Sheet1', index=False, header=True)
    # 再次读取数据
    sx_file = '筛选数据后的文件2.xlsx'
    sx_name = 'Sheet1'

    # 业务
    df_iptv = pd.read_excel(uploaded_file,
                       usecols='D,BJ,B',
                       header=0)
    # olt
    df_olt = pd.read_excel(uploaded_file,
                           usecols='D,BJ,AU',
                           header=0)
    # 此处为各部门参加问卷调查人数
    df_qingxi = pd.read_excel(sx_file,
                                    sheet_name=sx_name,
                                    usecols='A:P',
                                    header=0)

    # 多页面程序
    class CisAPP:
        def __init__(self):
            self.apps = []
            self.app_dict = {}

        def add_app(self, title, func):
            if title not in self.apps:
                self.apps.append(title)
                self.app_dict[title] = func

        def run(self):
            title = st.sidebar.radio(
                '菜单栏',
                self.apps,
                format_func=lambda title: str(title))
            self.app_dict[title]()

    def yem1():
        # 设置网页标题
        st.header('质差用户数据详情')
        # 展示导入的表格
        st.dataframe(df)
        # st.subheader('IPTV质差用户数据可视化')
        st.balloons()
        # streamlit的多重选择(选项数据)
        sjxz = df_iptv['时间'].unique().tolist()
        # streamlit的滑动条(年龄数据)
        cishu = df_iptv['卡顿次数'].unique().tolist()
        # 滑动条, 最大值、最小值、区间值
        cishu_sele = st.slider('卡顿次数:',
                                  min_value=min(cishu),
                                  max_value=max(cishu),
                                  value=(min(cishu), max(cishu)))
        # 多重选择, 默认全选
        sj_sele = st.multiselect('时间:',
                                              sjxz,
                                              default=sjxz)
        # 根据选择过滤数据
        mask = (df_iptv['卡顿次数'].between(*cishu_sele)) & (df_iptv['时间'].isin(sj_sele))
        numb_result = df_iptv[mask].shape[0]

        # 根据筛选条件, 得到有效数据
        st.markdown(f'*有效数据: {numb_result}*')
        # 根据选择分组数据
        df_gro1 = df_iptv[mask].groupby(by=['机顶盒业务账号']).count()[['卡顿次数']]
        df_gro1 = df_gro1.rename(columns={'卡顿次数': '次数'})
        df_gro1 = df_gro1.reset_index()

        bar1 = px.bar(df_gro1.head(15),
                           y='机顶盒业务账号',
                           x='次数',
                           text='次数',
                           color_discrete_sequence=['#F63366'] * len(df_gro1),
                           template='plotly_white',
                           orientation='h',
                           width=850,
                           height=700,
                           )
        st.plotly_chart(bar1)

        #交互式表格
        columns1, columns2 = st.columns(2)
        image = Image.open('wanfeng.jpg')
        columns1.image(image,
                   caption='业务帐号卡顿详情->',
                   width=500,
                   use_column_width=True)
        columns2.dataframe(df_iptv[mask], width=500)

        pie1 = px.pie(df_qingxi,
                           title='时间分布图',
                           values='时间重复次数',
                           names='时间',
                           height=600, )
        st.plotly_chart(pie1)

    def yem2():
        # streamlit的多重选择(选项数据)
        sj = df_olt['时间'].unique().tolist()
        # streamlit的滑动条(年龄数据)
        cishu = df_olt['卡顿次数'].unique().tolist()
        # 滑动条, 最大值、最小值、区间值
        kaduan_sele = st.slider('卡顿次数:',
                                  min_value=min(cishu),
                                  max_value=max(cishu),
                                  value=(min(cishu), max(cishu)))
        # 多重选择, 默认全选
        sj_sele = st.multiselect('时间:',
                                              sj,
                                              default=sj)
        # 根据选择过滤数据
        mask1 = (df_olt['卡顿次数'].between(*kaduan_sele)) & (df_olt['时间'].isin(sj_sele))
        numb_result = df_olt[mask1].shape[0]

        # 根据筛选条件, 得到有效数据
        st.markdown(f'*有效数据: {numb_result}*')
        # 根据选择分组数据
        df_gro2 = df_olt[mask1].groupby(by=['olt设备ip']).count()[['卡顿次数']]
        df_gro2 = df_gro2.rename(columns={'卡顿次数': '次数'})
        df_gro2 = df_gro2.reset_index()

        bar2 = px.bar(df_gro2.head(15),
                           y='olt设备ip',
                           x='次数',
                           text='次数',
                           color_discrete_sequence=['#F63366'] * len(df_gro2),
                           template='plotly_white',
                           orientation='h',
                           width=850,
                           height=700,
                           )
        st.plotly_chart(bar2)

        # 添加图片和交互式表格
        columns1, columns2 = st.columns(2)
        image = Image.open('riluo.jpg')
        columns1.image(image,
                   caption='riluo',
                   use_column_width=True)
        columns2.dataframe(df_olt[mask1], width=700)
        # 扇形图
        pie2 = px.pie(df_qingxi,
                           title='bras设备ip分布图',
                           values='bras设备ip重复次数',
                           names='bras设备ip',
                           height=600, )
        st.plotly_chart(pie2)

    def yem3():
        # 扇形图
        pie3 = px.pie(df_qingxi,
                           title='硬件型号分布图',
                           values='硬件型号重复次数',
                           names='硬件型号',
                           height=600, )
        st.plotly_chart(pie3)
        # 扇形图
        pie4 = px.pie(df_qingxi,
                           title='机顶盒软件版本号分布图',
                           values='机顶盒软件版本号重复次数',
                           names='机顶盒软件版本号',
                           height=600, )
        st.plotly_chart(pie4)

    def yem4():
        workb = load_workbook('筛选数据后的文件2.xlsx')
        wd = workb.active
        liks = []
        for i in range(1, wd.max_column + 1):            # 每列循环
            lk = 1
            for j in range(1, wd.max_row + 1):          # 每行循环
                sz = wd.cell(row=j, column=i).value
                if isinstance(sz, str):             # 中文占用多个字节，需要分开处理
                    lk1 = len(sz.encode('gbk'))
                else:
                    lk1 = len(str(sz))
                if lk < lk1:
                    lk = lk1
            liks.append(lk)

        # 第二步：设置列宽
        for i in range(1, wd.max_column + 1):
            k = get_column_letter(i)
            wd.column_dimensions[k].width = liks[i - 1] + 2  # 设置列宽，一般加两个字节宽度，可以根据实际情况灵活调整
        workb.close()
        workb.save('筛选数据后的文件2.xlsx')
        with open("筛选数据后的文件2.xlsx", "rb") as file:
                btn = st.download_button(label="下载TOP质差数据",
                                         data=file, file_name="TOP质差用户详情.xlsx")
        st.balloons()
    app = CisAPP()
    app.add_app("业务帐号时间分布情况", yem1)
    app.add_app("OLT、MSE分布趋势", yem2)
    app.add_app('机顶盒型号详情信息', yem3)
    app.add_app('TOP质差数据下载', yem4)
    app.run()


# 图片展示
def imageShow():
    image = Image.open("guidao.jpg")  # 在未导入excel时，展示图片
    st.image(image, clamp=False,
             channels="RGB", output_format="auto",use_column_width=True)


# 调用
head_set()
excel_upload()

# 如果文件已上传
if uploaded_file is not None:
    shuju()

# 否则展示图片
else:
    imageShow()